from openpyxl import Workbook, load_workbook 
import os
import time
from datetime import datetime, date, time
import datos_row

def slices(
    s,
    args=(0, 24, 31, 37, 43, 51, 58, 88,99, 107, 120, -1),
    wanted_columns = {1, 2, 3, 5, 7,8, 10}
    ):
    try:
        for column in wanted_columns:
            yield s[args[column]:args[column+1]].rstrip()
    except IndexError:
        return

def filtra_excel(
    ruta_entrada,
    linea,
    paradas,
    días
    ):
    ruta_entrada=os.getcwd()+datos_row.archivo
##    libro_entrada = load_workbook(ruta_entrada)
##    hoja_entrada = libro_entrada.active
    libro_salida = Workbook(write_only=True)
    hoja_salida = libro_salida.create_sheet()
    hoja_salida.append(
        ("Macro",
         "Línea",
         "Coche",
         "Parada",
         "Instante",
         " ",
         "Nombre"
         )
        )
    with open(ruta_entrada, "r", encoding="utf-8") as archivo_entrada:
        iter_archivo = iter(archivo_entrada)
        next(iter_archivo)
        next(iter_archivo)
        for raw_row in iter_archivo:
            try:
                row=list(slices(raw_row))
                row[1] = int(row[1])
                row[2] = int(row[2])
                row[3] = int(row[3])
                row[4] = datetime.strptime(row[4], "%Y-%m-%d").date()
                row[5] = datetime.strptime(row[5], "%H:%M:%S").time()
                if row and (row[1]==linea
                    and row[3] in paradas
                    and row[4].day in días):
                    hoja_salida.append(row)
            except ValueError:
                print("Row mala")
                print(row)
                continue
    libro_salida.save(
        os.getcwd() + "\\LINEA{linea}.xlsx".format(linea=linea))

def filtra_excel_coordinacion(
    nombre_archivo,
    ruta_entrada,
    lineas,
    paradas,
    días
    ):
    ruta_entrada=os.getcwd()+datos_row.archivo
##    libro_entrada = load_workbook(ruta_entrada)
##    hoja_entrada = libro_entrada.active
    libro_salida = Workbook(write_only=True)
    hoja_salida = libro_salida.create_sheet()
    hoja_salida.append(
        ("Macro",
         "Línea",
         "Coche",
         "Parada",
         "Instante",
         " ",
         "Nombre"
         )
        )
    with open(ruta_entrada, "r", encoding="utf-8") as archivo_entrada:
        iter_archivo = iter(archivo_entrada)
        next(iter_archivo)
        next(iter_archivo)
        for raw_row in iter_archivo:
            try:
                row=list(slices(raw_row))
                row[1] = int(row[1])
                row[2] = int(row[2])
                row[3] = int(row[3])
                row[4] = datetime.strptime(row[4], "%Y-%m-%d").date()
                row[5] = datetime.strptime(row[5], "%H:%M:%S").time()  
                if row and (row[1] in lineas
                    and row[3] in paradas
                    and row[4].day in días):
                    hoja_salida.append(row)
            except ValueError:
                print("Row mala")
                print(row)
                continue
    libro_salida.save(
        os.getcwd() + "\\{nombre_archivo}.xlsx".format(nombre_archivo=nombre_archivo))

