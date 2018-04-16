from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from datetime import timedelta, datetime
import os
import filtra_excel

class bar():
    
    hour = minute= second = None
    value = None
# open the file
ruta=os.getcwd()
nombre='LINEA17'
xlsx = (ruta+'/'+nombre+".xlsx")
xlsx_entrada =(ruta+"/"+'Libro1.xlsx')

def itera_vehiculo(vehiculo):
    for row in ws1:
        if row[2].value==vehiculo:
            yield row
filtra_excel.filtra_excel(xlsx_entrada,17,(443,465,10,44,394,400,423,415),(tuple(range(1,29))))   
wb = load_workbook(filename = xlsx)
ws1 = wb.active
ws3 = wb.create_sheet(title="Deporte")
ws4 = wb.create_sheet(title="La Gloria")
ws3.append(('coche','hora cabecera','El Mazo 4 caminos','San Fdo 66 Somo 118','null', 'duracion trayecto'))
ws4.append(('coche','hora cabecera','El Mazo 4 caminos','San Fdo 66 Somo 118','null', 'duracion trayecto'))

conjunto_vehiculos=set((row[2].value for row in ws1))  
for vehiculo in conjunto_vehiculos:
    fila_anterior=(bar(),bar(),bar(),bar(),bar(),bar(),bar())
    fila_2_anterior=(bar(),bar(),bar(),bar(),bar(),bar(),bar())
    tiempo1=tiempo2=tiempo3=entradaInterc=salidaInterc=cabecera=None
    gloria=deporte=False
    for row in itera_vehiculo(vehiculo):
        if row[3].value==10 and fila_2_anterior[3].value==443:
            if tiempo1 and cabecera and entradaInterc and tiempo2:
                nueva_fila=(vehiculo,datetime.combine(cabecera[4].value,cabecera[5].value),tiempo1,tiempo2,0)
                if gloria:
                    ws4.append(nueva_fila)
                elif deporte:
                    ws3.append(nueva_fila)
                gloria=deporte=False
                tiempo1=tiempo2=tiempo3=entradaInterc=salidaInterc=cabecera=None
            t1=row[5].value
            t2=fila_2_anterior[5].value
            dt1 = timedelta(hours=t1.hour, minutes=t1.minute, seconds=t1.second)
            dt2 = timedelta(hours=t2.hour, minutes=t2.minute, seconds=t2.second)
            if dt1-dt2 > timedelta(0):
                tiempo1=dt1-dt2
                entradaInterc=dt1
                cabecera=fila_anterior
        if row[3].value==400 and fila_2_anterior[3].value==44:
            t1=row[5].value
            t2=fila_2_anterior[5].value
            dt1 = timedelta(hours=t1.hour, minutes=t1.minute, seconds=t1.second)
            dt2 = timedelta(hours=t2.hour, minutes=t2.minute, seconds=t2.second)
            if dt1-dt2 > timedelta(0):
                if fila_anterior[3].value ==465 or fila_anterior[3].value ==394:
                    print('deporte')
                    deporte=True
                    gloria=False
                elif fila_anterior[3].value ==423 or fila_anterior[3].value ==415:
                    print('gloria')
                    gloria=True
                    deporte=False
                else:
                    gloria=deporte=False
                tiempo2=dt1-dt2
                salidaInterc=dt2
        fila_2_anterior=fila_anterior
        fila_anterior=row
        
    
wb.save(filename = ruta+'/'+nombre+"_parseada_febrero17.xlsx")    
