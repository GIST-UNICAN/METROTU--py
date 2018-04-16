
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from datetime import timedelta, datetime
import os
import mysql.connector
import contextlib
import bisect
import filtra_excel

    

class bar():
    
    hour = minute= second = None
    value = None
# open the file
query = ("SELECT Instante FROM `pasos_parada` where linea=100 and parada=516 ORDER BY `Instante` asc")

with contextlib.closing(mysql.connector.connect(user='root',
                                                password='madremia902',
                                                host='193.144.208.142',
                                                port='3306',
                                                database='autobuses')
                        ) as conexion:
    with contextlib.closing(conexion.cursor(buffered=True)) as cursor:
        cursor.execute(query)
        pasos_central=list((x[0] + timedelta(hours=1) for x in cursor))
        pasos_central.sort()
        print(pasos_central[-1])
        
ruta=os.getcwd()
nombre='COORDINACION_SARDINERO_BARRIOS'
xlsx = (ruta+"/"+nombre+'.xlsx')
xlsx_entrada =(ruta+"/"+'Libro1.xlsx')
filtra_excel.filtra_excel_coordinacion(nombre,xlsx_entrada,(20,9),(515,),(20,))        


def itera_linea(linea):
    for row in ws1:
        if row[1].value==linea:
            yield row
def check_cnx(fecha):
    llegada_central=pasos_central[bisect.bisect_right(pasos_central,fecha)-1] 
    espera=fecha-llegada_central
    if espera > timedelta(hours=1):
        return None
    else:
        return (0, espera,llegada_central) if espera > timedelta(minutes=6) else (1, espera,llegada_central)
        
        
wb = load_workbook(filename = xlsx)
ws1 = wb.active
ws3 = wb.create_sheet(title="CENTRO BARRIOS")
ws3.append(('linea','coche','hora_salida', 'llegada_central' ,'conexion', 'espera', 'dia','hora'))
conjunto_linea=set((row[1].value for row in ws1))  
for linea in conjunto_linea:
    fila_anterior=(bar(),bar(),bar(),bar(),bar(),bar(),bar())
    for row in itera_linea(linea):
        if fila_anterior[5].value: #row[1].value==3:
            t1=row[5].value
            t2=fila_anterior[5].value
            dt1 = timedelta(hours=t1.hour, minutes=t1.minute, seconds=t1.second)
            dt2 = timedelta(hours=t2.hour, minutes=t2.minute, seconds=t2.second)
            conexion=check_cnx(datetime.combine(row[4].value,row[5].value))
            if (dt1-dt2 < timedelta(minutes=3) or linea is 9) and conexion:
                if conexion[0] == 0:
                    fecha=datetime.combine(row[4].value,row[5].value)
                    nueva_fila=(linea,row[2].value,fecha,conexion[2],conexion[0],conexion[1],fecha.day,fecha.hour)
                    ws3.append(nueva_fila)
                else:
                    fecha=datetime.combine(row[4].value,row[5].value)
                    nueva_fila=(linea,row[2].value,fecha,conexion[2],conexion[0],conexion[1],fecha.day,fecha.hour)
                    ws3.append(nueva_fila)
        fila_anterior=row
    
wb.save(filename = ruta+"/"+nombre+'_parseada_dia 20.xlsx')    
