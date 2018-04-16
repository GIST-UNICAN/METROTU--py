from openpyxl import Workbook #load_workbook, 
import os
import time

def slices(
    s,
    args=(0, 24, 31, 37, 43, 51, 58, 88, 107, 120, -1),
    wanted_columns = {1, 2, 3, 5, 10, 11, 12}
    ):
    iter_args=iter(args)
##    position_0 = next(iter_args)
##    for column_index, position_1 in enumerate(iter_args):
##        if column_index in wanted_columns:
##            yield s[position_0:position_1].rstrip()
##        position_0 = position_1
    for column in wanted_columns:
        yield s[args[column]:args[column+1]

def filtra_excel(
    ruta_entrada,
    linea,
    paradas,
    días
    ):
##    libro_entrada = load_workbook(ruta_entrada)
##    hoja_entrada = libro_entrada.active
    libro_salida = Workbook(write_only=True)
    hoja_salida = libro_salida.active
    hoja_salida.append(
        ("Macro",
         "Línea",
         "Coche",
         "Parada",
         "Instante",
         " ",
         "Nombre"
         )
        )
    with open(ruta_entrada, "r") as archivo_entrada:
        for row in slices(raw_row):
            if (row[1].value==linea
                and row[3].value in paradas
                and row[4].value.day in días):
                hoja_salida.append(tuple(slices(x.value for x in row)))
    libro_salida.save(
        os.getcwd() + "\\LINEA{linea}.xlsx".format(linea=linea))

def filtra_excel_coordinacion(
    nombre_archivo,
    ruta_entrada,
    lineas,
    paradas,
    días
    ):
    libro_entrada = load_workbook(ruta_entrada)
    hoja_entrada = libro_entrada.active
##    print(hoja_entrada.title)
    libro_salida = Workbook()
    hoja_salida = libro_salida.active
    hoja_salida.append(
        ("Macro",
         "Línea",
         "Coche",
         "Parada",
         "Instante",
         " ",
         "Nombre"
         )
        )
    for row in slices(raw_row):
        if (row[1].value in lineas
            and row[3].value in paradas
            and row[4].value.day in días):
            hoja_salida.append(tuple((x.value for x in row)))
    libro_salida.save(
        os.getcwd() + "\\{nombre_archivo}.xlsx".format(nombre_archivo=nombre_archivo))

