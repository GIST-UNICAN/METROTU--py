from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from datetime import timedelta, datetime
import os

class bar():
    
    hour = minute= second = None
    value = None
# open the file
ruta=os.getcwd()
nombre='LINEA9'
xlsx = (ruta+"/"+nombre+'.xlsx')

def itera_vehiculo(vehiculo):
    for row in ws1:
        if row[2].value==vehiculo:
            yield row
    
wb = load_workbook(filename = xlsx)
ws1 = wb.active
ws3 = wb.create_sheet(title="tv repuente sardinero")
ws4 = wb.create_sheet(title="tv sardinero cueto")
ws5 = wb.create_sheet(title="tv cueto sardinero")
ws6 = wb.create_sheet(title="tv sardinero repuente")
ws3.append(('coche','hora cabecera', 'duracion trayecto'))
ws4.append(('coche','hora cabecera', 'duracion trayecto'))
ws5.append(('coche','hora cabecera', 'duracion trayecto'))
ws6.append(('coche','hora cabecera', 'duracion trayecto'))
conjunto_vehiculos=set((row[2].value for row in ws1))  
for vehiculo in conjunto_vehiculos:
    fila_anterior=(bar(),bar(),bar(),bar(),bar(),bar(),bar())
    for row in itera_vehiculo(vehiculo):
        if row[3].value==511 and fila_anterior[3].value==158:
            t1=row[5].value
            t2=fila_anterior[5].value
            dt1 = timedelta(hours=t1.hour, minutes=t1.minute, seconds=t1.second)
            dt2 = timedelta(hours=t2.hour, minutes=t2.minute, seconds=t2.second)
            nueva_fila=(vehiculo,datetime.combine(fila_anterior[4].value,fila_anterior[5].value),dt1-dt2)
            if dt1-dt2 > timedelta(0):
                ws3.append(nueva_fila)
        if row[3].value==251 and fila_anterior[3].value==511:
            t1=row[5].value
            t2=fila_anterior[5].value
            dt1 = timedelta(hours=t1.hour, minutes=t1.minute, seconds=t1.second)
            dt2 = timedelta(hours=t2.hour, minutes=t2.minute, seconds=t2.second)
            nueva_fila=(vehiculo,datetime.combine(fila_anterior[4].value,fila_anterior[5].value),dt1-dt2)
            if dt1-dt2 > timedelta(0):
                ws4.append(nueva_fila)
        if row[3].value==515 and fila_anterior[3].value==251:
            t1=row[5].value
            t2=fila_anterior[5].value
            dt1 = timedelta(hours=t1.hour, minutes=t1.minute, seconds=t1.second)
            dt2 = timedelta(hours=t2.hour, minutes=t2.minute, seconds=t2.second)
            nueva_fila=(vehiculo,datetime.combine(fila_anterior[4].value,fila_anterior[5].value),dt1-dt2)
            if dt1-dt2 > timedelta(0):
                ws5.append(nueva_fila)
        if row[3].value==511 and fila_anterior[3].value==158:
            t1=row[5].value
            t2=fila_anterior[5].value
            dt1 = timedelta(hours=t1.hour, minutes=t1.minute, seconds=t1.second)
            dt2 = timedelta(hours=t2.hour, minutes=t2.minute, seconds=t2.second)
            nueva_fila=(vehiculo,datetime.combine(fila_anterior[4].value,fila_anterior[5].value),dt1-dt2)
            if dt1-dt2 > timedelta(0):
                ws6.append(nueva_fila)
        fila_anterior=row
    
wb.save(filename = ruta+"/"+nombre+'_parseada.xlsx')    
