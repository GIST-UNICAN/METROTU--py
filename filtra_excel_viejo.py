from openpyxl import Workbook #load_workbook, 
import os
import time

def slices(
    s,
    iter_args=iter((0, 24, 31, 37, 43, 51, 58, 88, 107, 120, -1)),
    wanted_columns = {1, 2, 3, 5, 10, 11, 12}
    ):
##    print(type(s))
    position_0 = next(iter_args)
    for column_index, position_1 in enumerate(iter_args):
##        print("pos0", position_0)
##        print("pos1", position_1)
        if column_index in wanted_columns:
            print("devuelve", s[position_0:position_1].rstrip())
            yield s[position_0:position_1].rstrip()
        position_0 = position_1
        

def filtra_excel(
    ruta_entrada,
    linea,
    paradas,
    días
    ):
    ruta_entrada=os.getcwd()+'/enero.rpt'
##    libro_entrada = load_workbook(ruta_entrada)
##    hoja_entrada = libro_entrada.active
    libro_salida = Workbook(write_only=True)
    hoja_salida = libro_salida.create_sheet()
    hoja_salida.append(
        ("Macro",
         "Línea",
         "Coche",
         "Parada",
         "Instante",
         " ",
         "Nombre"
         )
        )
##    with open(ruta_entrada, "r") as archivo_entrada:
##        for raw_row in archivo_entrada:
##            print(raw_row)
##            time.sleep(1)
    with open(ruta_entrada, "r", encoding="utf-8") as archivo_entrada:
        for raw_row in archivo_entrada:
            print(raw_row)
            row=tuple(slices(raw_row))
            print(row)
            time.sleep(0.5)
##        for row in (tuple(slices(raw_row)) for raw_row in archivo_entrada):
##            print(raw_row)
##            print(row)
            if row and (row[1]==linea
                and row[3] in paradas
                and row[4].day in días):
                hoja_salida.append(row)
    libro_salida.save(
        os.getcwd() + "\\LINEA{linea}.xlsx".format(linea=linea))

def filtra_excel_coordinacion(
    nombre_archivo,
    ruta_entrada,
    lineas,
    paradas,
    días
    ):
    libro_entrada = load_workbook(ruta_entrada)
    hoja_entrada = libro_entrada.active
##    print(hoja_entrada.title)
    libro_salida = Workbook(write_only=True)
    hoja_salida = libro_salida.create_sheet()
    hoja_salida.append(
        ("Macro",
         "Línea",
         "Coche",
         "Parada",
         "Instante",
         " ",
         "Nombre"
         )
        )
    for row in (tuple(slices(raw_row)) for raw_row in archivo_entrada):
        if (row[1] in lineas
            and row[3] in paradas
            and row[4].day in días):
            hoja_salida.append(row)
    libro_salida.save(
        os.getcwd() + "\\{nombre_archivo}.xlsx".format(nombre_archivo=nombre_archivo))

