from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from datetime import timedelta, datetime
import os
import filtra_excel

class bar():
    
    hour = minute= second = None
    value = None
# open the file
ruta=os.getcwd()
nombre='LINEA62'
xlsx = (ruta+'/'+nombre+".xlsx")
xlsx_entrada =(ruta+"/"+'Libro1.xlsx')

def itera_vehiculo(vehiculo):
    for row in ws1:
        if row[2].value==vehiculo:
            yield row
filtra_excel.filtra_excel(xlsx_entrada,62,(183,212),(tuple(range(1,29))))   
wb = load_workbook(filename = xlsx)
ws1 = wb.active
ws3 = wb.create_sheet(title="tv")
ws3.append(('coche','hora cabecera','Polideportivo Avda Deporte 6','P.Ayto  Avda Deporte 15','null', 'duracion trayecto'))
conjunto_vehiculos=set((row[2].value for row in ws1))  
for vehiculo in conjunto_vehiculos:
    fila_anterior=(bar(),bar(),bar(),bar(),bar(),bar(),bar())
    tiempo1=tiempo2=tiempo3=entradaInterc=salidaInterc=cabecera=None
    for row in itera_vehiculo(vehiculo):
        if row[3].value==212 and fila_anterior[3].value==183:
            if tiempo1 and cabecera and entradaInterc:
                nueva_fila=(vehiculo,datetime.combine(cabecera[4].value,cabecera[5].value),tiempo1,tiempo2,0)
                ws3.append(nueva_fila)
                tiempo1=tiempo2=tiempo3=entradaInterc=salidaInterc=cabecera=None
            t1=row[5].value
            t2=fila_anterior[5].value
            dt1 = timedelta(hours=t1.hour, minutes=t1.minute, seconds=t1.second)
            dt2 = timedelta(hours=t2.hour, minutes=t2.minute, seconds=t2.second)
            if dt1-dt2 > timedelta(0):
                tiempo1=dt1-dt2
                entradaInterc=dt1
                cabecera=fila_anterior
        if row[3].value==181 and fila_anterior[3].value==41:
            t1=row[5].value
            t2=fila_anterior[5].value
            dt1 = timedelta(hours=t1.hour, minutes=t1.minute, seconds=t1.second)
            dt2 = timedelta(hours=t2.hour, minutes=t2.minute, seconds=t2.second)
            if dt1-dt2 > timedelta(0):
                tiempo2=dt1-dt2
                salidaInterc=dt2
        fila_anterior=row
    
wb.save(filename = ruta+'/'+nombre+"_parseada_febrero17.xlsx")    
