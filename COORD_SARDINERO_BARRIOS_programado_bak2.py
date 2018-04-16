
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from datetime import timedelta, datetime
import os
##import mysql.connector
import contextlib
import bisect
import filtra_excel
from collections import namedtuple
import MySQLdb
from bisect import bisect
from functools import partial
from itertools import repeat, starmap

class bar():

    hour = minute= second = None
    value = None

Números_de_paradas = namedtuple(
    "Números_de_paradas",
    (
        "Sardinero",
        "Sardinero_1",
        "Sardinero_2",
        "Valdecilla",
        "Avda_Valdecilla",
        "San_Fernando",
        "Vega_Lamera",
        "Cajo2",
        "TQuevedo22",
        "Albericia18")
    )
números_de_paradas = Números_de_paradas(516, 511, 515, 509, 512)

def entrecomilla(s):
    return "".join("'", s, "'")

def filtra_central_sardinero(llegada):
    return hora_salida_sardinero

def filtra_central_valdecilla(llegada):
    return hora_salida_valdecilla

##def comprueba(salida_row_anterior,salida, minutos_salida=(6,21,36,51), tiempo_hasta_parada_control=timedelta(minutes=1, seconds=30), margen_tiempo=6):
##    punto_insercion=bisect(minutos_salida,(salida-tiempo_hasta_parada_control).minute)
##    diferencia_minutos=(salida-tiempo_hasta_parada_control).minute-minutos_salida[punto_insercion-1]
##    if diferencia_minutos<0:
##        diferencia_minutos+=60
##        hora=salida.hour-1
##    else:
##        hora=salida.hour
##
##    if diferencia_minutos>margen_tiempo:
##        return salida_row_anterior
##    else:
##        return salida.replace(hour=hora,minute=minutos_salida[punto_insercion-1], second=0)



class Comprobador():
    def __init__(
        self,
        minutos_salida,
        tiempo_hasta_parada_control,
        margen_tiempo,
        horamin=(
            (13, 45)
            (13, 45)
        ):
        self.minutos_salida = minutos_salida
        self.tiempo_hasta_parada_control = tiempo_hasta_parada_control
        self.margen_tiempo = margen_tiempo

    def __call__(self, salida_row_anterior, salida):
        punto_insercion=bisect(self.minutos_salida,(salida-self.tiempo_hasta_parada_control).minute)
        diferencia_minutos=(salida-self.tiempo_hasta_parada_control).minute-self.minutos_salida[punto_insercion-1]
        if diferencia_minutos<0:
            diferencia_minutos+=60
            hora=salida.hour-1
        else:
            hora=salida.hour

        if diferencia_minutos>self.margen_tiempo:
            return salida_row_anterior
        else:
            return salida.replace(hour=hora,minute=self.minutos_salida[punto_insercion-1], second=0)

(
    comprueba_3,
    comprueba_13,
    comprueba_17,
    comprueba_100_sardi,
    comprueba_100_valde
    )= starmap(
        Comprobador,
        (
            ((6,21,36,51),timedelta(minutes=1, seconds=30), 6),
            ###((6,21,36,51),timedelta(minutes=1, seconds=30), margen_tiempo=6),
            ###((6,21,36,51),timedelta(minutes=1, seconds=30), margen_tiempo=6),
            ###((6,21,36,51),timedelta(minutes=1, seconds=30), margen_tiempo=6),
            ####((6,21,36,51),timedelta(minutes=1, seconds=30), margen_tiempo=6),
            ((8,23,38,53),timedelta(minutes=1, seconds=30), 6)
            )
        )
##    comprueba_3 = partial(comprueba,

# open the file
actual=datetime.now()
fecha_inicio=datetime(actual.year,actual.month,actual.day,6,0,0)
fecha_fin=datetime(actual.year,actual.month,actual.day,20,0,0)
paradas_query = " or ".join(f"parada={parada}" for parada in números_de_paradas)
query = (f"SELECT Instante FROM `pasos_parada` where Instante between '{fecha_inicio}' and '{fecha_fin}' "
         f"and ({paradas_query}) "
         "ORDER BY Linea, Coche, Instante asc"
         )
print(query)
##query = (f"SELECT fecha_hora FROM `pasos_parada` where Instante between {fecha_inicio} and {fecha_fin} "
##         "and (parada=516 or parada=511 or parada=515 or parada=509 or parada=512 or parada=11 or parada=171) "
##         "ORDER BY `Linea`, Coche, Instante"
##         )

with contextlib.closing(MySQLdb.connect(user='root',
                                                password='madremia902',
                                                host='193.144.208.142',
                                                port=3306,
                                                database='autobuses')
                        ) as conexion:
    with contextlib.closing(conexion.cursor()) as cursor:
        cursor.execute(query)
        datos_row=tuple(cursor)

# lineas 91  repuente sardinero repuente y 92 cueto sardinero cueto

llegadas_sardinero={100:list(),91:list(),20:list(), 92: list()}
llegadas_valdecilla={100:list(),3:list(),13:list(),17:list()}
salidas_sardinero={100:list(),91:list(),20:list(), 92: list()}
salidas_valdecilla={100:list(),3:list(),13:list(),17:list()}

for row in datos_row:
    linea=row[0]
    instante=row[4]
    parada=row[3]
    if linea==100:
        if parada==paradas.Avda_Valdecilla:
            llegadas_valdecilla[100].append(instante)
        if parada==paradas.San_Fernando:
            salidas_valdecilla[100].append(comprueba_100_valde(linea_anterior[4],instante))
    if linea==100:
        if parada==paradas.Sardinero:
            llegadas_sardinero[100].append(instante)
        if parada==paradas.Vega_Lamera:
            salidas_sardinero[100].append(comprueba_100_valde(linea_anterior[4],instante))
    if linea==3:
        if parada==paradas.Valdecilla:
            llegadas_valdecilla[3].append(instante)
        if parada==paradas.Cajo2:
            salidas_valdecilla[3].append(comprueba_3(linea_anterior[4],instante))
    if linea==13:
        if parada==paradas.Valdecilla:
            llegadas_valdecilla[13].append(instante)
        if parada==paradas.TQuevedo22:
            salidas_valdecilla[13].append(comprueba_13(linea_anterior[4],instante))
    if linea==17:
        if parada==paradas.Valdecilla:
            llegadas_valdecilla[17].append(instante)
        if parada==paradas.Albericia18:
            salidas_valdecilla[17].append(comprueba_17(linea_anterior[4],instante))
    if linea==20:
        if parada==paradas.Sardinero1:
            llegadas_sardinero[20].append(instante)
        if parada==paradas.Sardinero2:
            salidas_sardinero[20].append(instante)
    if linea==9:
        if parada==paradas.Sardinero1:
            llegadas_sardinero[91].append(instante)
            salidas_sardinero[92].append(instante)
        if parada==paradas.Sardinero2:
            llegadas_sardinero[92].append(instante)
            salidas_sardinero[91].append(instante)
    
    linea_anterior=row

##ruta=os.getcwd()
##nombre='COORDINACION_SARDINERO_BARRIOS'
##xlsx = (ruta+"/"+nombre+'.xlsx')
##xlsx_entrada =(ruta+"/"+'Libro1.xlsx')
##filtra_excel.filtra_excel_coordinacion(nombre,xlsx_entrada,(20,9),(515,),(16,))
##
##
##def itera_linea(linea):
##    for row in ws1:
##        if row[1].value==linea:
##            yield row
##def check_cnx(fecha):
##    llegada_central=pasos_central[bisect.bisect_right(pasos_central,fecha)-1]
##    espera=fecha-llegada_central
##    if espera > timedelta(hours=1):
##        return None
##    else:
##        return (0, espera,llegada_central) if espera > timedelta(minutes=6) else (1, espera,llegada_central)
##
##
##wb = load_workbook(filename = xlsx)
##ws1 = wb.active
##ws3 = wb.create_sheet(title="CENTRO BARRIOS")
##ws3.append(('linea','coche','hora_salida', 'llegada_central' ,'conexion', 'espera', 'dia','hora'))
##conjunto_linea=set((row[1].value for row in ws1))
##for linea in conjunto_linea:
##    fila_anterior=(bar(),bar(),bar(),bar(),bar(),bar(),bar())
##    for row in itera_linea(linea):
##        if fila_anterior[5].value: #row[1].value==3:
##            t1=row[5].value
##            t2=fila_anterior[5].value
##            dt1 = timedelta(hours=t1.hour, minutes=t1.minute, seconds=t1.second)
##            dt2 = timedelta(hours=t2.hour, minutes=t2.minute, seconds=t2.second)
##            conexion=check_cnx(datetime.combine(row[4].value,row[5].value))
##            if (dt1-dt2 < timedelta(minutes=3) or linea is 9) and conexion:
##                if conexion[0] == 0:
##                    fecha=datetime.combine(row[4].value,row[5].value)
##                    nueva_fila=(linea,row[2].value,fecha,conexion[2],conexion[0],conexion[1],fecha.day,fecha.hour)
##                    ws3.append(nueva_fila)
##                else:
##                    fecha=datetime.combine(row[4].value,row[5].value)
##                    nueva_fila=(linea,row[2].value,fecha,conexion[2],conexion[0],conexion[1],fecha.day,fecha.hour)
##                    ws3.append(nueva_fila)
##        fila_anterior=row
##
##wb.save(filename = ruta+"/"+nombre+'_parseada_dia 16.xlsx')
