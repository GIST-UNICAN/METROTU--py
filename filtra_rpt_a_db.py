from openpyxl import Workbook, load_workbook 
import os
import time
from datetime import datetime, date
import contextlib
import MySQLdb

def slices(
    s,
    #args=(0, 31, 37, 45, 51, 75, 87, 92, 98, 177,196,201,203,205,-1),
    #wanted_columns = {1, 3, 5, 7, 9, 10, 11,12}
    args=(0, 24, 30, 36, 43,104,111, 119,138,143,150,152,-1),
    wanted_columns = {1,2,3,4,5,7,9,10}
    ):
    try:
        for column in wanted_columns:
            yield s[args[column]:args[column+1]].rstrip()
    except IndexError:
        return

def filtra_excel(
    ruta_entrada
    ):
    query_append=list()
    c_rr=0
    c_rm=0
    #ruta_entrada=os.getcwd()+datos_row.archivo
##    libro_entrada = load_workbook(ruta_entrada)
##    hoja_entrada = libro_entrada.active
    libro_salida = Workbook()
    hoja_salida = libro_salida.create_sheet()
    hoja_salida.append(
        ("Linea",
         "Coche",
         "Parada",
         "Nombre",
         "Viaje",
         "Instante",
         "Psuben",
         "Fecha"
         )
        )
    with open(ruta_entrada, "r", encoding="utf-8") as archivo_entrada:
        iter_archivo = iter(archivo_entrada)
        next(iter_archivo)
        next(iter_archivo)
        
        for raw_row in iter_archivo:
            try:
                c_rr+=1
                row=list(slices(raw_row))
                row[0] = int(row[0])
                row[1] = int(row[1])
                row[2] = int(row[2])
                row[3] = str(row[3])
                row[4] = int(row[4])
                if row[5]== 'NULL':
                    continue
                row[5] = datetime.strptime(str(row[5]), "%Y-%m-%d %H:%M:%S")
                row[6] =999 if row[6]== 'NULL' else  int(row[6]) 
                row[7] = datetime.strptime(str(row[5]), "%Y-%m-%d %H:%M:%S").replace(hour=0, minute=0, second=0)
                #insertamos en la bd
                query_append.append("({},{},{},{},'{}','{}',{},'{}')".format(
                        row[0],row[1],row[4],row[2],row[5],row[3],row[6],row[7]))
                
                
                hoja_salida.append(row)
            except ValueError as e:
                print(e)
                c_rm+=1
                print("Row mala")
                print(row)
                continue
    libro_salida.save(
        os.getcwd() + "\\LINEA{linea}.xlsx".format(linea='112'))
    querie="INSERT INTO `pasos_parada` (`Linea`, `Coche`, `Viaje`, `Parada`, `Instante`, `Nombre`, `PSuben`, `Fecha`) VALUES " +str(query_append)[1:-1].replace('"',"")
    try:
        with contextlib.closing(MySQLdb.connect(
                host='193.144.208.142',
                user="root",
                passwd="madremia902",
                db="autobuses")) as cnx:
            with contextlib.closing(cnx.cursor()) as crs:
                crs.execute(querie)
                cnx.commit()
    except Exception as e:
        print('sql error' + str(e))
    print(c_rr)
    print(c_rm)

filtra_excel(
    r"D:\Users\Andres\OneDrive\OneDrive - Universidad de Cantabria\Recordar GIST - VARIOS\datos_tus\16.rpt"
    )

