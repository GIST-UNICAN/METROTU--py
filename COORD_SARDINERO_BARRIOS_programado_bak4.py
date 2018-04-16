
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from datetime import timedelta, datetime,time
import os
##import mysql.connector
import contextlib
import bisect
import filtra_excel
from collections import namedtuple
import MySQLdb
from bisect import bisect
from functools import partial
from itertools import starmap, groupby
import horarios
from collections import defaultdict
import pandas as pd
import numpy as np
import traceback
from functools import reduce


class bar():
    
    hour = minute= second = None
    value = None

Números_de_paradas = namedtuple(
    "Números_de_paradas",
    (
        "Sardinero",
        "Sardinero1",
        "Sardinero2",
        "Valdecilla",
        "Avda_Valdecilla",
        "San_Fernando",
        "Vega_Lamera",
        "Cajo2",
        "TQuevedo22",
        "Albericia18")
    )
números_de_paradas = Números_de_paradas(516, 511, 515, 509, 512, 11, 171, 78, 46, 390)

def entrecomilla(s):
    return "".join("'", s, "'")

def datetime2time(fecha):
    return time(fecha.hour,fecha.minute,fecha.second)

def comparatiempos(t1, t2): #devuelve da diferencia ABSOLUTA
    t1=t1.hour*3600+t1.minute*60+t1.second
    t2=t2.hour*3600+t2.minute*60+t2.second
    diferencia= abs(t1-t2)
    horas, resto=divmod(diferencia,3600)
    minutos, segundos =divmod(resto,60)
    return time(horas,minutos,segundos)

def restatiempos(t1, t2): 
    t1=t1.hour*3600+t1.minute*60+t1.second
    t2=t2.hour*3600+t2.minute*60+t2.second
    diferencia= t1-t2
    horas, resto=divmod(diferencia,3600)
    minutos, segundos =divmod(resto,60)
    return time(horas,minutos,segundos)

def time2datetime(lista):
    dia=datetime.now().day
    mes=datetime.now().month
    año=datetime.now().year
    return tuple((datetime(año,mes,dia,parada.hour,parada.minute,parada.second) for parada in lista))

def ordena_diccionario(dic):
    for k,v in dic.items():
        dic[k] = tuple(x for x,y in groupby(sorted(v)))
    return dic

def filtra_llegadas_central(lista):
    devuelve_lista=list()
    elemento_anterior=datetime(1900,1,1,0,0,0)
    for elemento in lista:
        if elemento - elemento_anterior > timedelta(minutes=2):
            devuelve_lista.append(elemento)
        elemento_anterior=elemento
    return devuelve_lista

class Comprobador():
    
        

    def __init__(
        self,
        horario_linea,
        tiempo_hasta_parada_control,
        margen_tiempo
        ):
        self.horario_linea = time2datetime(horario_linea)
        self.tiempo_hasta_parada_control = tiempo_hasta_parada_control
        self.margen_tiempo = margen_tiempo

    def __call__(self, salida_row_anterior, salida, cero_segundos=time()):
##        print('---------')
##        print(salida_row_anterior)
##        print(salida)
        punto_insercion=bisect(self.horario_linea,(salida-self.tiempo_hasta_parada_control))
        diferencia=(salida-self.tiempo_hasta_parada_control)-self.horario_linea[punto_insercion]
        diferencia2=(salida-self.tiempo_hasta_parada_control)-self.horario_linea[punto_insercion-1]
##        print(diferencia)
##        print(self.horario_linea[punto_insercion])
##        print(self.horario_linea[punto_insercion-1])
##        print(abs(diferencia)>self.margen_tiempo)
        if abs(diferencia)>self.margen_tiempo and abs(diferencia2)>self.margen_tiempo: #se perdio la conexion y sale a la hora de paso por valdecilla
##            print(salida_row_anterior)
            return salida_row_anterior
        elif abs(diferencia)>abs(diferencia2):
            #volvemos a comprobar la salida anterior con la teorica para ver cual es mayor y nos quedamos con la mayor
##            print(self.horario_linea[punto_insercion-1] if (self.horario_linea[punto_insercion-1]>salida_row_anterior) else salida_row_anterior)
            return self.horario_linea[punto_insercion-1] if (self.horario_linea[punto_insercion-1]>salida_row_anterior) else salida_row_anterior
        else:
##            print(self.horario_linea[punto_insercion] if (self.horario_linea[punto_insercion]>salida_row_anterior) else salida_row_anterior)
            return self.horario_linea[punto_insercion] if (self.horario_linea[punto_insercion]>salida_row_anterior) else salida_row_anterior
            #devolvemos la salida según el horario programado

(
    comprueba_3,
    comprueba_13,
    comprueba_17,
    comprueba_100_sardi,
    comprueba_100_valde
    )= starmap(
        Comprobador,
        (
            (horarios.paradas_3,timedelta(minutes=1, seconds=30), timedelta(minutes=6)),
            (horarios.paradas_13,timedelta(minutes=2, seconds=0), timedelta(minutes=6)),
            (horarios.paradas_17,timedelta(minutes=1, seconds=30), timedelta(minutes=6)),
            (horarios.paradas_C_Int_Sardinero,timedelta(minutes=2, seconds=00), timedelta(minutes=6)),
            (horarios.paradas_C_Int_Valdecilla,timedelta(minutes=3, seconds=00), timedelta(minutes=6))
            )
        )

# open the file
actual=datetime.now()
fecha_inicio=datetime(actual.year,actual.month,actual.day,6,0,0)
fecha_fin=datetime(actual.year,actual.month,actual.day,20,0,0)
paradas_query = " or ".join(f"parada={parada}" for parada in números_de_paradas)
query = (f"SELECT * FROM `pasos_parada` where Instante between '{fecha_inicio}' and '{fecha_fin}' "
         f"and ({paradas_query}) "
         "ORDER BY Linea, Coche, Instante asc"
         )
print(query)

with contextlib.closing(MySQLdb.connect(user='root',
                                                password='madremia902',
                                                host='193.144.208.142',
                                                port=3306,
                                                database='autobuses')
                        ) as conexion:
    with contextlib.closing(conexion.cursor()) as cursor:
        cursor.execute(query)
        datos_row=tuple(cursor)

# lineas 91  repuente sardinero repuente y 92 cueto sardinero cueto

llegadas_sardinero={100:list(),91:list(),20:list(), 92: list()}
llegadas_valdecilla={100:list(),3:list(),13:list(),17:list()}
salidas_sardinero={100:list(),91:list(),20:list(), 92: list()}
salidas_valdecilla={100:list(),3:list(),13:list(),17:list()}           

for linea_bus, rows in groupby(datos_row, lambda fila: (fila[0], fila[1])):
    linea=linea_bus[0]
    
    for row in rows:
        instante=row[4]
        parada=row[3]
        if linea==100:
            if parada==números_de_paradas.Avda_Valdecilla:
                llegadas_valdecilla[100].append(instante)
            if parada==números_de_paradas.San_Fernando:
                salidas_valdecilla[100].append(comprueba_100_valde(linea_anterior[4],instante))
        if linea==100:
            if parada==números_de_paradas.Sardinero:
                llegadas_sardinero[100].append(instante)
            if parada==números_de_paradas.Vega_Lamera:
                salidas_sardinero[100].append(comprueba_100_sardi(linea_anterior[4],instante))
        if linea==3:
            if parada==números_de_paradas.Valdecilla:
                llegadas_valdecilla[3].append(instante)
            if parada==números_de_paradas.Cajo2:
                salidas_valdecilla[3].append(comprueba_3(linea_anterior[4],instante))
        if linea==13:
            if parada==números_de_paradas.Valdecilla:
                llegadas_valdecilla[13].append(instante)
            if parada==números_de_paradas.TQuevedo22:
                salidas_valdecilla[13].append(comprueba_13(linea_anterior[4],instante))
        if linea==17:
            if parada==números_de_paradas.Valdecilla:
                llegadas_valdecilla[17].append(instante)
            if parada==números_de_paradas.Albericia18:
                salidas_valdecilla[17].append(comprueba_17(linea_anterior[4],instante))
        if linea==20:
            if parada==números_de_paradas.Sardinero1:
                llegadas_sardinero[20].append(instante)
            if parada==números_de_paradas.Sardinero2:
                salidas_sardinero[20].append(instante)
        if linea==9:
            if parada==números_de_paradas.Sardinero1:
                llegadas_sardinero[91].append(instante)
                salidas_sardinero[92].append(instante)
            if parada==números_de_paradas.Sardinero2:
                llegadas_sardinero[92].append(instante)
                salidas_sardinero[91].append(instante)
        
        linea_anterior=row

lineas_acaban_sardinero = (91, 92, 20)
lineas_acaban_valdecilla = (3, 13, 17)


#ordenamos los diccionarios con llegadas y salidas
(
    llegadas_sardinero,
    llegadas_valdecilla,
    salidas_sardinero,
    salidas_valdecilla
    ) = map(ordena_diccionario,
            (
                llegadas_sardinero,
                llegadas_valdecilla,
                salidas_sardinero,
                salidas_valdecilla
                )
            )
# generamos unas listas con las corresondencias para cada intercambiador

resultados = {
    "Valdecilla_barrios": [],
    "Barrios_Valdecilla": [],
    "Sardinero_barrios": [],
    "Barrios_Sardinero": []
    }

llegadas_valdecilla[100]=filtra_llegadas_central(llegadas_valdecilla[100])
for instante_llegada_central in llegadas_valdecilla[100]:
    for linea in lineas_acaban_valdecilla:
        posición_salida_linea = bisect(salidas_valdecilla[linea], instante_llegada_central)
        instante_salida_linea = salidas_valdecilla[linea][posición_salida_linea]
        tiempo_espera = instante_salida_linea - instante_llegada_central
        resultados["Valdecilla_barrios"].append((linea,
                                                instante_salida_linea,
                                                instante_llegada_central,
                                                tiempo_espera)
                                                )

for linea in lineas_acaban_valdecilla:
    for instante_llegada in llegadas_valdecilla[linea]:
        try:
            posición_salida_central = bisect(salidas_valdecilla[100], instante_llegada)
            instante_salida_central = salidas_valdecilla[100][posición_salida_central]
            tiempo_espera = instante_salida_central - instante_llegada
            if tiempo_espera<timedelta(hours=2):
                resultados["Barrios_Valdecilla"].append((linea,
                                                         instante_llegada,
                                                         instante_salida_central,
                                                         tiempo_espera))
        except IndexError:
            #print(traceback.format_exc())
            print(instante_llegada, ' ', linea)
            pass
        
for linea in lineas_acaban_sardinero:
    for instante_salida in salidas_sardinero[linea]:
        posición_llegada_central = bisect(llegadas_sardinero[100], instante_salida)-1
        instante_llegada_central = llegadas_sardinero[100][posición_llegada_central]
        tiempo_espera = instante_salida - instante_llegada_central
        if posición_llegada_central != -1:
            resultados["Sardinero_barrios"].append((linea,
                                                    instante_salida,
                                                    instante_llegada_central,
                                                    tiempo_espera)
                                                    )


    for instante_llegada in llegadas_sardinero[linea]:
        try:
            posición_salida_central = bisect(salidas_sardinero[100], instante_llegada)
            instante_salida_central = salidas_sardinero[100][posición_salida_central]
            tiempo_espera = instante_salida_central - instante_llegada
            if tiempo_espera<timedelta(hours=2):
                resultados["Barrios_Sardinero"].append((linea,
                                                         instante_llegada,
                                                         instante_salida_central,
                                                         tiempo_espera))
        except IndexError:
            print(instante_llegada, ' ', linea)
            pass
        
def media_espera(objeto):
    return reduce(timedelta.__add__,objeto)/len(objeto)

# agregamos los datos con pandas :)
dataframe_valdecilla_barrios=pd.DataFrame(resultados['Valdecilla_barrios'],
                                          columns=['linea','salida','llegada_central','espera'])
table_valdecilla_barrios=pd.pivot_table(
    dataframe_valdecilla_barrios,
    values=['espera'],
    index=['linea'],
    aggfunc=media_espera)


##llegadas_sardinero={100:list(),91:list(),20:list(), 92: list()}
##llegadas_valdecilla={100:list(),3:list(),13:list(),17:list()}
##salidas_sardinero={100:list(),91:list(),20:list(), 92: list()}
##salidas_valdecilla={100:list(),3:list(),13:list(),17:list()}
##ruta=os.getcwd()
##nombre='COORDINACION_SARDINERO_BARRIOS'
##xlsx = (ruta+"/"+nombre+'.xlsx')
##xlsx_entrada =(ruta+"/"+'Libro1.xlsx')
##filtra_excel.filtra_excel_coordinacion(nombre,xlsx_entrada,(20,9),(515,),(16,))
##
##
##def itera_linea(linea):
##    for row in ws1:
##        if row[1].value==linea:
##            yield row
##def check_cnx(fecha):
##    llegada_central=pasos_central[bisect.bisect_right(pasos_central,fecha)-1]
##    espera=fecha-llegada_central
##    if espera > timedelta(hours=1):
##        return None
##    else:
##        return (0, espera,llegada_central) if espera > timedelta(minutes=6) else (1, espera,llegada_central)
##
##
##wb = load_workbook(filename = xlsx)
##ws1 = wb.active
##ws3 = wb.create_sheet(title="CENTRO BARRIOS")
##ws3.append(('linea','coche','hora_salida', 'llegada_central' ,'conexion', 'espera', 'dia','hora'))
##conjunto_linea=set((row[1].value for row in ws1))
##for linea in conjunto_linea:
##    fila_anterior=(bar(),bar(),bar(),bar(),bar(),bar(),bar())
##    for row in itera_linea(linea):
##        if fila_anterior[5].value: #row[1].value==3:
##            t1=row[5].value
##            t2=fila_anterior[5].value
##            dt1 = timedelta(hours=t1.hour, minutes=t1.minute, seconds=t1.second)
##            dt2 = timedelta(hours=t2.hour, minutes=t2.minute, seconds=t2.second)
##            conexion=check_cnx(datetime.combine(row[4].value,row[5].value))
##            if (dt1-dt2 < timedelta(minutes=3) or linea is 9) and conexion:
##                if conexion[0] == 0:
##                    fecha=datetime.combine(row[4].value,row[5].value)
##                    nueva_fila=(linea,row[2].value,fecha,conexion[2],conexion[0],conexion[1],fecha.day,fecha.hour)
##                    ws3.append(nueva_fila)
##                else:
##                    fecha=datetime.combine(row[4].value,row[5].value)
##                    nueva_fila=(linea,row[2].value,fecha,conexion[2],conexion[0],conexion[1],fecha.day,fecha.hour)
##                    ws3.append(nueva_fila)
##        fila_anterior=row
##
##wb.save(filename = ruta+"/"+nombre+'_parseada_dia 16.xlsx')
