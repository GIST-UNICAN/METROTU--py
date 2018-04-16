
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from datetime import timedelta, datetime
import os
import filtra_excel

class bar():
    
    hour = minute= second = None
    value = None
# open the file
ruta=os.getcwd()
nombre='COORDINACION_BARRIOS_VALDECILLA'
entrada='Libro1'
xlsx = (ruta+"/"+nombre+'.xlsx')
xlsx_entrada = (ruta+"/"+entrada+'.xlsx')


def itera_linea(linea):
    for row in ws1:
        if row[1].value==linea:
            yield row
def check_minute(minuto,fecha):
    if minuto>=8 and minuto<23:
        salida_central=fecha.replace(minute=23, second=0, microsecond=0)
        return (0, abs(minuto - 8),salida_central) if abs(minuto-8) < 11 else (1, abs(minuto - 8),salida_central)
    if minuto>=23 and minuto<38:
        salida_central=fecha.replace(minute=38, second=0, microsecond=0)
        return (0, abs(minuto - 23),salida_central) if abs(minuto-23) < 11 else (1, abs(minuto - 23),salida_central)
    if minuto>=38 and minuto<53:
        salida_central=fecha.replace(minute=53, second=0, microsecond=0)
        return (0, abs(minuto - 38),salida_central) if abs(minuto-38) < 11 else (1, abs(minuto - 38),salida_central)
    if minuto>=53 and minuto<60:
        salida_central=fecha.replace(hour=fecha.hour+1,minute=8, second=0, microsecond=0)
        return (0, abs(minuto - 53),salida_central) if abs(minuto-53) < 11 else (1, abs(minuto - 53),salida_central)
    if minuto>=00 and minuto<8:
        salida_central=fecha.replace(minute=8, second=0, microsecond=0)
        return(0, abs(minuto - 0+7),salida_central) if abs(minuto-0+7) < 11 else (1, abs(minuto - 0+7),salida_central)
filtra_excel.filtra_excel_coordinacion(nombre,xlsx_entrada,(17,13,3,),(512,),(20,))    
wb = load_workbook(filename = xlsx)
ws1 = wb.active
ws3 = wb.create_sheet(title="CENTRO BARRIOS")
ws3.append(('linea','coche','llegada_bus','salida_central','conexion', 'espera','dia','hora'))
conjunto_linea=set((row[1].value for row in ws1))  
for linea in conjunto_linea:
    fila_anterior=(bar(),bar(),bar(),bar(),bar(),bar(),bar())
    tiempo_anterior=datetime(1900,1,1,0,0,0)
    for row in itera_linea(linea):
        if fila_anterior[5].value: #row[1].value==3:
            t1=row[5].value
            t2=fila_anterior[5].value
            dt1 = timedelta(hours=t1.hour, minutes=t1.minute, seconds=t1.second)
            dt2 = timedelta(hours=t2.hour, minutes=t2.minute, seconds=t2.second)
            fecha=datetime.combine(row[4].value,row[5].value)
            conexion=check_minute(t1.minute,fecha)
            if fecha-tiempo_anterior>timedelta(minutes=10):
                if conexion[0] == 0:
                    salida_central=fecha
                    fecha=datetime.combine(row[4].value,row[5].value)
                    nueva_fila=(linea,row[2].value,fecha,conexion[2],conexion[0],15-conexion[1],fecha.day,fecha.hour)
                    tiempo_anterior=fecha
                    ws3.append(nueva_fila)
                else:
                    fecha=datetime.combine(row[4].value,row[5].value)
                    nueva_fila=(linea,row[2].value,fecha,conexion[2],conexion[0],15-conexion[1],fecha.day,fecha.hour)
                    tiempo_anterior=fecha
                    ws3.append(nueva_fila)

        fila_anterior=row
    
wb.save(filename = ruta+"/"+nombre+'_parseada_dia 20.xlsx')    
